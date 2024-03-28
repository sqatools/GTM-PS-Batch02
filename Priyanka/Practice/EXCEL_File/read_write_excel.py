"""
pip install openpyxl
"""

import openpyxl

def read_excel_data(filename, row_num, col_num, sheet_name):
    wb = openpyxl.load_workbook(filename)
    excel_sheet = wb[f'{sheet_name}']
    #print(excel_sheet)
    cell_obj = excel_sheet.cell(row=row_num, column=col_num)
    print(cell_obj.value)


#read_excel_data("Excel_data.xlsx", row_num=1, col_num=1,sheet_name="Sheet1" )

def read_excel_data_method2(filename, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename)
    # provide excel sheet parameter name with f string formatting
    excel_sheet = wb[f'{sheet_name}']
    #print(excel_sheet)
    cell_obj = excel_sheet[f'{cell_name}']
    print(cell_obj.value)


#read_excel_data_method2("Excel_data.xlsx", sheet_name='Sheet2', cell_name="A5")

def write_excel_data_method(filename, sheet_name, cell_name, new_val):
    wb = openpyxl.load_workbook(filename)
    excel_sheet = wb[sheet_name]
    cell_obj = excel_sheet[f'{cell_name}']
    cell_obj.value = new_val
    print(cell_obj.value)
    wb.save(filename)


for i in range (1,6):
    write_excel_data_method("Excel_data.xlsx", sheet_name='Sheet2', cell_name=f"B{i}", new_val=f"{i}")


