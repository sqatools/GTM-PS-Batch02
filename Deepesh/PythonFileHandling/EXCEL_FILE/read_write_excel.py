"""
pip install openpyxl
"""
import openpyxl

def read_excel_data(filename, row_num, col_num, sheet_name):
    wb = openpyxl.load_workbook(filename)
    excel_sheet = wb[f'{sheet_name}']
    #print(excel_sheet)
    cell_obj = excel_sheet.cell(row=row_num, column=col_num)
    print(cell_obj.value)


#read_excel_data("read_excel_data.xlsx", row_num=1, col_num=1, )
#read_excel_data("read_excel_data.xlsx", row_num=2, col_num=1)

#for i in range(1, 6):
#    read_excel_data("read_excel_data.xlsx", row_num=i, col_num=1, sheet_name="Sheet2")


def read_excel_data_method2(filename, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename)
    # provide excel sheet parameter name with f string formatting
    excel_sheet = wb[f'{sheet_name}']
    #print(excel_sheet)
    cell_obj = excel_sheet[f'{cell_name}']
    print(cell_obj.value)


#read_excel_data_method2("read_excel_data.xlsx", sheet_name='Sheet1', cell_name="B5")

# for i in range(1, 6):
#     read_excel_data_method2("read_excel_data.xlsx", sheet_name='Sheet1', cell_name=f"B{i}")


# Create new excel file and add content
def write_content_to_file(filename, sheet_name, cell_name, cell_value):
    wb = openpyxl.Workbook()
    sheet_obj = wb.active
    cell = sheet_obj[f'{cell_name}']
    cell.value = cell_value
    wb.save(filename)


write_content_to_file("read_excel_data3.xlsx", "Sheet1", "A1", "Sanjay")

# add content to existing excel file
def write_content_to_file_without_overwrite(filename, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filename)
    sheet_obj = wb[sheet_name]
    cell = sheet_obj[f'{cell_name}']
    cell.value = cell_value
    wb.save(filename)

#write_content_to_file_without_overwrite("read_excel_data.xlsx", "Sheet2", "A2", "Modi Ji")

# for i in range(1, 10):
#     write_content_to_file_without_overwrite("read_excel_data.xlsx", "Sheet2", f"A{i}", f"{535534+i}")









