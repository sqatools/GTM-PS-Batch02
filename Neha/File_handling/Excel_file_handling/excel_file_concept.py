import openpyxl


# def read_excel_data(filename, row_num, col_num, sheet_name):
#     wb = openpyxl.load_workbook(filename)
#     excel_sheet = wb[f'{sheet_name}']
#     #print(excel_sheet)
#     cell_obj = excel_sheet.cell(row=row_num, column=col_num)
#     print(cell_obj.value)

#write
# def write_content_to_file(filename, sheet_name, cell_name, cell_value):
#     wb = openpyxl.Workbook()
#     sheet_obj = wb.active
#     cell = sheet_obj[f'{cell_name}']
#     cell.value = cell_value
#     wb.save(filename)
#
#
# write_content_to_file("read_excel_data1.xlsx", "Sheet1", "A1", "Neha")

# add content to existing excel file
def write_content_to_file_without_overwrite(filename, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filename)
    sheet_obj = wb[sheet_name]
    cell = sheet_obj[f'{cell_name}']
    cell.value = cell_value
    wb.save(filename)

write_content_to_file_without_overwrite("read_excel_data1.xlsx", "Sheet2", "A2", "Peeyush")

# for i in range(1, 10):
#     write_content_to_file_without_overwrite("read_excel_data.xlsx", "Sheet2", f"A{i}", f"{535534+i}")