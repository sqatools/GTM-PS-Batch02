import openpyxl

def write_excel_data(filename, sheet_name, Cell_name, Cell_value):
    wb = openpyxl.Workbook()
    excel_sheet_obj = wb.active
    cell = excel_sheet_obj[f'{Cell_name}']
    cell.value = Cell_value
    # Cell_value = cell.value
    print(Cell_value)

write_excel_data("read_excel_data2.xlsx", sheet_name="Sheet",Cell_name= 'B1',Cell_value='Neha')


# def excel_read_data(filename,row_num,Col_num,sheet_name):
#     wb = openpyxl.load_workbook(filename)
#     excel_sheet = wb['Sheet2']
#     cell_obj = excel_sheet.cell(row= row_num, column = Col_num)
#     print(cell_obj.value)
#
# excel_read_data("read_excel_data.xlsx",row_num = 4 ,Col_num= 1,sheet_name=2)


#excel_read_data_with_cellname
# def excel_read_data_with_cellname(filename , sheet_name, Cell_num ):
#     wb = openpyxl.load_workbook(filename)
#     excel_sheet = wb[f'{sheet_name}']
#     cell_obj = excel_sheet[f'{Cell_num}']
#     print(cell_obj.value)
#
# excel_read_data_with_cellname("read_excel_data.xlsx", sheet_name="Sheet", Cell_num="A2")